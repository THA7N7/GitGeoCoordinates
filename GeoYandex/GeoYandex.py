#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import sys
import requests
from bs4 import BeautifulSoup
import openpyxl
from os.path import join, abspath
import time


def print_description_text():
    '''
        ==========================================================================
        GeoPosition from Yandex.maps ver. 1.0.7

        Программа позволяет получить Широту и Долготу с Яндекс.Карты по Адресу.
        Программа работает с файлом «AddressFile.xlsx».
        Заголовки(!) [A1] == «Адрес» [B1] == «Широта» [C1] == «Долгота».
        В колонке A начиная со второй строки адреса без пропущенных строк.
        Желательно после адреса, через запятую, указать название объекта.
        Колонки B и C начиная со второй строки заполняет программа (поверх данных)

        tha7n7@gmail.com freeware ©2021
        ==========================================================================
    '''
    pass


def print_version_text():
    '''
        ========================================
        GeoPosition from Yandex.maps ver. 1.0.7
        tha7n7@gmail.com freeware ©2021
        ========================================
    '''
    pass


def site_parse(req_region, i):
    '''Geoparsing Yandex.Maps'''
    headers = {'acsept': '*/*',
               'user-agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                             'AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/75.0.3770.142 Safari/537.36'}
    base_url = 'https://yandex.ru/maps/213/moscow/?&text=' + req_region
    session = requests.Session()
    request = session.get(base_url, headers=headers)

    if request.status_code == 200:
        soup = BeautifulSoup(request.content, 'lxml')
        try:
            val = soup.find('script', attrs={'type': 'application/json'}).text.strip()
            coordinates = val[val.find("displayCoordinates", 1, len(val)) + 21:]
            coordinates = coordinates[0:coordinates.find("],", 0, len(coordinates))]
            val = coordinates.split(',')
            val.reverse()
            print(str(i + 1) + '. lat:' + str(val[0]) + ' lon:' + str(val[1]))
        except ValueError:
            val = []
            print('NoData')
        return val


def main():
    xlsx_file = join('.', 'AddressFile.xlsx')
    xlsx_file = abspath(xlsx_file)
    start_time = time.time()
    try:
        wb = openpyxl.load_workbook(filename=xlsx_file, read_only=False, data_only=True)
        sh = wb.active
        start_row = 2
        end_row = 2
        r = 1
        a = sh.cell(row=1, column=1).value
        b = sh.cell(row=1, column=2).value
        c = sh.cell(row=1, column=3).value
        req_list = []
        if a + b + c != 'АдресШиротаДолгота':
            print(print_description_text.__doc__)
            print("Неверный формат файла")
            return False
        else:
            try:
                wb.save(xlsx_file)
            except PermissionError:
                print(print_description_text.__doc__)
                print("Файл закрыт для записи")
                return False
        v = a
        while v != None:
            v = sh.cell(row=r, column=1).value
            end_row = r
            r += 1

        for i in range(start_row, end_row):
            req_list.append([(str(sh.cell(row=i, column=1).value))])

        for i in range(0, len(req_list)):
            add_url = req_list[i][0]
            text_answer = site_parse(add_url, i)
            if text_answer[0][2] != '.' or text_answer[1][2] != '.':
                add_url = ''
                text_answer = ['NoData', 'NoData']
            try:
                sh.cell(row=2 + i, column=2).value, sh.cell(row=2 + i, column=3).value = text_answer[0], text_answer[1]
            except Exception:
                pass

        wb.save(xlsx_file)
        print()
        print(print_version_text.__doc__)
        print("Все адреса обработаны. Результат в файле AddressFile.xlsx")
        elapsed_time = time.time() - start_time
        print('Elapsed time : {} sec'.format(round(elapsed_time, 2)))
        wb.close()
        return True
    except FileNotFoundError:
        print(print_description_text.__doc__)
        print("AddressFile.xlsx не найден")
        return False


if __name__ == '__main__':
    print()
    main()
    print()
    input("Press any key to continue...")
