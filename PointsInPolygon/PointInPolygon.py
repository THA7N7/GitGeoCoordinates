#!/usr/bin/env python
# -*- coding: UTF-8 -*-
import openpyxl
from os.path import join, abspath
from bs4 import BeautifulSoup


def print_description_text():
    '''
            ====================================================================
            SearchPointsInPolygon ver. 1.0.3

            Программа ищет Точки находящиеся внутри Регионов.
            Программа работает с файлом «PointsPost.xlsx».
            Заголовки(!) [A1] == «Индекс» [B1] == «Широта» [C1] == «Долгота».
            В колонке A начиная со второй строки Точки без пропущенных строк.
            Регионы поиска в файле «Polygons.kml». Результат записывается в файл
            «InsidePolygon.xlsx» (нужен пустой файл с таким именем)

            tha7n7@gmail.com freeware ©2021
            =====================================================================
        '''
    pass


def print_version_text():
    '''
        =================================
        SearchPointsInPolygon ver. 1.0.3
        tha7n7@gmail.com freeware ©2021
        =================================
    '''
    pass


def is_point_inside_polygon(x: float, y: float, poly) -> bool:
    """Determine if the point is in the polygon.
    Args:
      x -- The x coordinates of point.
      y -- The y coordinates of point.
      poly -- a list of tuples [(x, y), (x, y), ...]
    Returns:
      True if the point is in the polygon.
    """
    num = len(poly)
    i = 0
    j = num - 1
    c = False
    for i in range(num):
        if ((poly[i][1] > y) != (poly[j][1] > y)) and \
                (x < poly[i][0] + (poly[j][0] - poly[i][0]) * (y - poly[i][1]) /
                 (poly[j][1] - poly[i][1])):
            c = not c
        j = i
    return c


def process_coodinates_string(str):
    polygon_str = (' '.join(str.split())).split(' ')
    polygon_tup = []
    for s in polygon_str:
        if len(s.split(',')) == 1:
            p = tuple(map(float, s.split(',')[::-1]))
        else:
            p = tuple(map(float, s.split(',')[:2][::-1]))
        polygon_tup.append(p)
    return polygon_tup


def open_xlsx(name):
    xlsx_file = join('.', name)
    xlsx_file = abspath(xlsx_file)
    try:
        return openpyxl.load_workbook(filename=xlsx_file, read_only=False, data_only=True)
    except FileNotFoundError:
        print(print_description_text.__doc__)
        print("{} не найден".format(name))
        print()
        input("Press any key to continue...")
        quit()


def main():
    wb_points = open_xlsx('PointsPost.xlsx')
    sh = wb_points.active
    start_row = 2
    end_row = 2
    r = 1
    a = sh.cell(row=1, column=1).value
    b = sh.cell(row=1, column=2).value
    c = sh.cell(row=1, column=3).value
    point_list = []
    if a + b + c != 'ИндексШиротаДолгота':
        print(print_description_text.__doc__)
        print("Неверный формат файла")
        return False
    else:
        try:
            wb_points.save('PointsPost.xlsx')
        except PermissionError:
            print(print_description_text.__doc__)
            print("Файл закрыт для записи")
            return False

    v = a
    while v != None:
        v = sh.cell(row=r, column=1).value
        end_row = r
        r += 1

    polygon = []
    kml_file = join('.', 'Polygons.kml')
    with open(kml_file, 'r', encoding="utf-8") as f:
        soup = BeautifulSoup(f, 'xml')
        for coords in soup.find_all(('coordinates')):
            polygon.append(process_coodinates_string(coords.string))

    for i in range(start_row, end_row):
        point_list.append((int(sh.cell(row=i, column=1).value),
                           float(sh.cell(row=i, column=2).value),
                           float(sh.cell(row=i, column=3).value)))

    points_inside = []
    for i in range(0, len(point_list)):
        for j in range(0, len(polygon)):
            if is_point_inside_polygon(point_list[i][1], point_list[i][2], polygon[j]):
                points_inside.append(point_list[i][0])
                print(str(len(points_inside)) + ". " + str(points_inside[len(points_inside) - 1]))
                break

    if len(points_inside) > 0:
        wb_points_inside = open_xlsx('InsidePolygon.xlsx')
        sh = wb_points_inside.active
        sh.delete_cols(1)
        for i in range(0, len(points_inside)):
            sh.cell(row=1 + i, column=1).value = points_inside[i]
            wb_points_inside.save('InsidePolygon.xlsx')
        wb_points_inside.close()
    else:
        print("Точки внутри регионов не найдены")
    print(print_version_text().__doc__)
    wb_points.close()


if __name__ == '__main__':
    print()
    main()
    print()
    input("Press any key to continue...")
